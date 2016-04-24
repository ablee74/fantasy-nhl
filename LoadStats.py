#Loads stats for each player. Taken from excel database
#to player class objects
import openpyxl
from Player import *
wb = openpyxl.load_workbook('Player Database.xlsx')


def getPlayer(name, season):
    """This will pull player info from excel Database and put it in the
    player object - Player.py"""
    global wb
    ws = wb.get_sheet_by_name(season)
    notFound = True
    r = 5       #Names start at row 5
    while notFound and ws.cell(row = r, column = 1).value != None:
        if ws.cell(row = r, column = 1).value == name:
            notFound = False
        else:
            r = r + 1

    c = 2
    notFound = True
    while notFound and ws.cell(row = 4, column = c).value != None:
        if ws.cell(row = 4, column = c).value == "Team":
            notFound = False
        else:
            c = c + 1

    p = Player(name, ws.cell(row = r, column = c).value)

    traitL = ["G","Assists","TOI", "Shifts", "Games", "Shots on Goal",
              "Takeaways", "Giveaways", "Blocks", "Hits", "Faceoffs",
              "Won", "PIMs"]
    toAdd = []
    for trait in traitL:
        c = 2
        notFound = True
        while notFound and ws.cell(row = 4, column = c).value != None:
            if ws.cell(row = 4, column = c).value == trait:
                notFound = False
                toAdd.append(ws.cell(row = r, column = c).value)
            else:
                c = c + 1

    toAdd = blankConvert(toAdd)
    
    p.goals = int(toAdd[0])
    p.assists = int(toAdd[1])
    y = noComma(toAdd[2])
    p.TOI = int(y)
    y = noComma(toAdd[3])
    p.shifts = int(y)
    p.gamesPlayed = int(toAdd[4])
    p.SOG = int(toAdd[5])
    p.takeaways = int(toAdd[6])
    p.giveaways = int(toAdd[7])
    p.blocks = int(toAdd[8])
    p.hits = int(toAdd[9])
    p.faceoffs = int(toAdd[10])
    p.faceoffWins = int(toAdd[11])
    p.PIM = int(toAdd[12])
    
    return p # end line

def getHistory(season, name, stat):
    """This will read player stats by game spreadsheets to track stats per
    game in a season -- To be added to Player Object"""
    wb = openpyxl.load_workbook(season + '.xlsx')
    ws = wb.get_sheet_by_name(name)
    notFound = True
    c = 1   #First column in spreadsheet
    while notFound and ws.cell(row = 4, column = c).value != None:
        if ws.cell(row = 4, column = c).value == stat:
            notFound = False
        else:
            c = c + 1
    hL = []
    r = 5
    while ws.cell(row = r, column = c).value != None:
        hL.append(ws.cell(row = r, column = c).value)
        r = r + 1
    return hL
    
def noComma(x):
    """x is a string that represents a number, it has commas every 3
    digits"""
    if len(x)>3:
        y = x[:-4] + x[-3:]
    else:
        y = x

    return y

def blankConvert(L):
    """L is a list that represents player stats, changing BLANK on
    excel sheet to an integer aka 0"""
    for i in range(len(L)):
        if L[i] == 'BLANK':
            L[i] = '0'

    return L

def getRoster(team, season):
    """This will pull all players from a certain team to evaluate
    performance against another team"""
    global wb
    ws = wb.get_sheet_by_name(season)
    notFound = True
    r = 5   #teams start at row 5
    while notFound and ws.cell(row = r, column = 2).value != None:
        if ws.cell(row = r, column = 2).value == team:
            notFound = False
        else:
            r = r + 1

    roster = []

    while ws.cell(row = r, column = 2).value == team:
        p = getPlayer(ws.cell(row = r, column = 1).value, season)
        roster.append(p)
        r = r + 1

    return roster
    
        
            
        
    


    
    

    
    
    


